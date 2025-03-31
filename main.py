from flask import Flask, redirect, render_template, request, send_file, url_for, make_response
import logging
from extractors.wwr import WWRJobSearch
from extractors.wanted_job_search import WantedJobSearch
from extractors.job_data import JobData
from extractors.job_data_wwr import JobDataWWR
import csv
from datetime import datetime
import io
from utils.logger import setup_logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 로거 설정
logger = setup_logger(__name__)

app = Flask("JobScraper", static_folder='static', static_url_path='/static')

# 검색 결과를 저장할 인메모리 데이터베이스
db = {}

def create_csv_response(jobs, filename):
    """CSV 응답 생성 헬퍼 함수"""
    try:
        # CSV 버퍼 생성
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        
        # 헤더 결정
        if jobs:
            if len(jobs[0]) == 5:  # WWR 데이터
                headers = ["Position", "Company", "Location", "Salary", "Link"]
            else:  # Wanted 데이터
                headers = ["Position", "Company", "Reward", "Link"]
            
            # 헤더와 데이터 쓰기
            writer.writerow(headers)
            for job in jobs:
                writer.writerow(job)
            
            # 응답 생성
            response = make_response(csv_buffer.getvalue())
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-type"] = "text/csv; charset=utf-8"
            
            return response
        else:
            raise ValueError("No jobs data available")
            
    except Exception as e:
        logger.error(f"CSV 생성 중 오류 발생: {str(e)}")
        raise

def create_excel_response(jobs, filename):
    """Excel 응답 생성 헬퍼 함수"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        
        # 헤더 결정
        if jobs:
            if len(jobs[0]) == 5:  # WWR 데이터
                headers = ["Position", "Company", "Location", "Salary", "Link"]
            else:  # Wanted 데이터
                headers = ["Position", "Company", "Reward", "Link"]
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 헤더 작성
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 작성
            for row, job in enumerate(jobs, 2):
                for col, value in enumerate(job, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 메모리에 저장
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # 응답 생성
            response = make_response(excel_buffer.getvalue())
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            return response
        else:
            raise ValueError("No jobs data available")
            
    except Exception as e:
        logger.error(f"Excel 생성 중 오류 발생: {str(e)}")
        raise

@app.route("/")
def home():
    """홈페이지 렌더링"""
    return render_template("home.html")

@app.route("/search")
def search():
    """검색 키워드에 따른 구직 정보 검색 및 결과 페이지 렌더링"""
    keyword = request.args.get("keyword")
    if not keyword or keyword == "":
        logger.warning("빈 키워드로 검색 시도")
        return redirect("/")
        
    keywords = keyword.split(",")
    formatted_keywords = ", ".join([k.strip() for k in keywords])
    all_jobs = []
    
    for keyword in keywords:
        keyword = keyword.strip()
        if keyword in db:
            logger.info(f"캐시에서 키워드 '{keyword}'에 대한 결과 사용")
            jobs = db[keyword]
        else:
            logger.info(f"키워드 '{keyword}'에 대한 새 검색 수행")
            try:
                logger.info(f"1. Wanted 사이트 스크래핑 시작: 키워드 '{keyword}'")
                wanted = WantedJobSearch(headless=False).scrape_keyword(keyword)
                logger.info(f"Wanted 검색 결과: {len(wanted)}개 작업 찾음")
                
                logger.info(f"2. WWR 사이트 스크래핑 시작: 키워드 '{keyword}'")
                wwr = WWRJobSearch().scrape_keyword(keyword)
                logger.info(f"WWR 검색 결과: {len(wwr)}개 작업 찾음")
                
                # 결과 순서 변경: WWR가 상단에 표시되도록 함
                jobs = wwr + wanted
                db[keyword] = jobs
                logger.info(f"총 {len(jobs)}개 작업을 찾았습니다 (WWR: {len(wwr)}, Wanted: {len(wanted)})")
            except Exception as e:
                logger.error(f"스크래핑 중 오류 발생: {e}")
                jobs = []
        
        all_jobs.extend(jobs)
    
    search_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return render_template(
        "search.html",
        keyword=formatted_keywords,
        jobs=all_jobs,
        jobs_count=len(all_jobs),
        time=search_time
    )

@app.route("/export")
def export():
    """저장된 검색 결과 페이지 렌더링 또는 파일 내보내기"""
    keyword = request.args.get("keyword")
    file_type = request.args.get("type", "csv")  # 기본값은 csv
    
    # 파일 다운로드 요청인 경우
    if keyword and request.args.get("download") == "true":
        if keyword not in db:
            logger.warning(f"내보내기 요청: 키워드 '{keyword}'에 대한 데이터 없음")
            return redirect(f"/search?keyword={keyword}")
        
        jobs = db[keyword]
        if not jobs:
            logger.warning(f"내보내기 요청: 키워드 '{keyword}'에 대한 작업 목록이 비어 있음")
            return redirect(f"/search?keyword={keyword}")
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if file_type == "excel":
                filename = f"jobs_{keyword}_{timestamp}.xlsx"
                response = create_excel_response(jobs, filename)
                logger.info(f"Excel 파일 내보내기 성공: {filename}, {len(jobs)}개 작업")
            else:
                filename = f"jobs_{keyword}_{timestamp}.csv"
                response = create_csv_response(jobs, filename)
                logger.info(f"CSV 파일 내보내기 성공: {filename}, {len(jobs)}개 작업")
            return response
        except Exception as e:
            logger.error(f"파일 생성 중 오류 발생: {e}")
            return redirect(f"/search?keyword={keyword}")
    
    # 저장된 검색 결과 페이지 렌더링
    return render_template(
        "export.html",
        db=db,
        selected_keyword=keyword,
        jobs=db.get(keyword, []) if keyword else []
    )

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8080, debug=True)
